#!/usr/bin/env python3
"""
Data I/O and Serialization
==========================
Demonstrates data reading, writing, and serialization with various formats.

JSON: orjson, ujson, simplejson (fast JSON libraries)
XML: lxml, xmltodict, ElementTree
YAML: PyYAML
Excel: openpyxl, xlrd
Parquet: pyarrow, fastparquet
HDF5: h5py, tables (PyTables)
HTTP: requests, httpx
"""

import json
import numpy as np
import pandas as pd
import tempfile
import os

# =============================================================================
# JSON Libraries Comparison
# =============================================================================
print("=" * 60)
print("JSON Libraries")
print("=" * 60)

import orjson
import ujson
import simplejson

# Sample data
sample_data = {
    "users": [
        {"id": 1, "name": "Alice", "scores": [95, 87, 92]},
        {"id": 2, "name": "Bob", "scores": [78, 85, 90]},
    ],
    "metadata": {"version": "1.0", "generated": "2024-01-01"},
    "count": 2,
}

# Standard library json
json_str = json.dumps(sample_data, indent=2)
print("Standard json library:")
print(json_str[:100] + "...")

# orjson (fastest, returns bytes)
orjson_bytes = orjson.dumps(sample_data)
print(f"\norjson output (bytes): {orjson_bytes[:50]}...")

# Decode and pretty print
orjson_str = orjson.dumps(sample_data, option=orjson.OPT_INDENT_2).decode("utf-8")
print(f"orjson pretty: {orjson_str[:100]}...")

# ujson
ujson_str = ujson.dumps(sample_data, indent=2)
print(f"\nujson output: {ujson_str[:100]}...")

# simplejson
simplejson_str = simplejson.dumps(sample_data, indent=2)
print(f"\nsimplejson output: {simplejson_str[:100]}...")

# Parsing
parsed = orjson.loads(orjson_bytes)
print(f"\nParsed data keys: {list(parsed.keys())}")

# =============================================================================
# XML Libraries
# =============================================================================
print("\n" + "=" * 60)
print("XML Libraries")
print("=" * 60)

from lxml import etree
import xmltodict
import xml.etree.ElementTree as ET

# Sample XML
xml_string = """<?xml version="1.0" encoding="UTF-8"?>
<catalog>
    <book id="1">
        <title>Python Programming</title>
        <author>John Doe</author>
        <price>29.99</price>
    </book>
    <book id="2">
        <title>Data Science Handbook</title>
        <author>Jane Smith</author>
        <price>39.99</price>
    </book>
</catalog>
"""

# lxml parsing
print("lxml parsing:")
root = etree.fromstring(xml_string.encode())
for book in root.findall("book"):
    title = book.find("title").text
    print(f"  Book: {title}")

# xmltodict (XML to dictionary)
print("\nxmltodict conversion:")
xml_dict = xmltodict.parse(xml_string)
print(f"  Parsed type: {type(xml_dict)}")
print(f"  Books: {len(xml_dict['catalog']['book'])}")

# Convert dict back to XML
xml_output = xmltodict.unparse(xml_dict, pretty=True)
print(f"  Re-serialized length: {len(xml_output)} chars")

# Standard library ElementTree
print("\nElementTree parsing:")
et_root = ET.fromstring(xml_string)
for book in et_root.findall("book"):
    book_id = book.get("id")
    author = book.find("author").text
    print(f"  Book {book_id}: by {author}")

# =============================================================================
# YAML
# =============================================================================
print("\n" + "=" * 60)
print("YAML Processing")
print("=" * 60)

import yaml

yaml_data = """
database:
  host: localhost
  port: 5432
  credentials:
    username: admin
    password: secret

servers:
  - name: web1
    ip: 192.168.1.1
  - name: web2
    ip: 192.168.1.2

features:
  enabled: true
  max_connections: 100
"""

# Parse YAML
config = yaml.safe_load(yaml_data)
print("Parsed YAML config:")
print(f"  Database host: {config['database']['host']}")
print(f"  Number of servers: {len(config['servers'])}")
print(f"  Features enabled: {config['features']['enabled']}")

# Dump to YAML
output_yaml = yaml.dump(config, default_flow_style=False)
print(f"\nDumped YAML:\n{output_yaml[:200]}...")

# =============================================================================
# Excel Files
# =============================================================================
print("\n" + "=" * 60)
print("Excel File Handling")
print("=" * 60)

import openpyxl

# Create sample DataFrame
df = pd.DataFrame(
    {
        "Name": ["Alice", "Bob", "Charlie"],
        "Age": [25, 30, 35],
        "Salary": [50000, 60000, 70000],
    }
)

# Write to Excel
with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
    excel_path = f.name

df.to_excel(excel_path, sheet_name="Employees", index=False)
print(f"Written DataFrame to Excel: {excel_path}")

# Read back
df_read = pd.read_excel(excel_path, sheet_name="Employees")
print(f"Read back DataFrame shape: {df_read.shape}")
print(df_read)

# Using openpyxl directly
wb = openpyxl.load_workbook(excel_path)
ws = wb.active
print(f"\nSheet name: {ws.title}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

os.unlink(excel_path)

# =============================================================================
# Parquet Files
# =============================================================================
print("\n" + "=" * 60)
print("Parquet File Handling")
print("=" * 60)

import pyarrow as pa
import pyarrow.parquet as pq
import fastparquet

# Create larger DataFrame
np.random.seed(42)
df_large = pd.DataFrame(
    {
        "id": range(10000),
        "value": np.random.randn(10000),
        "category": np.random.choice(["A", "B", "C"], 10000),
        "date": pd.date_range("2024-01-01", periods=10000, freq="h"),
    }
)

# Write with PyArrow
with tempfile.NamedTemporaryFile(suffix=".parquet", delete=False) as f:
    parquet_path = f.name

df_large.to_parquet(parquet_path, engine="pyarrow", compression="snappy")
parquet_size = os.path.getsize(parquet_path)
print(f"Written Parquet file: {parquet_size / 1024:.2f} KB")

# Read with PyArrow
df_parquet = pd.read_parquet(parquet_path, engine="pyarrow")
print(f"Read back DataFrame shape: {df_parquet.shape}")

# Read specific columns
df_partial = pd.read_parquet(parquet_path, columns=["id", "value"])
print(f"Partial read (2 columns): {df_partial.shape}")

# Parquet metadata
parquet_file = pq.read_table(parquet_path)
print(f"\nParquet schema:")
print(parquet_file.schema)

# FastParquet
df_fp = fastparquet.ParquetFile(parquet_path).to_pandas()
print(f"\nFastparquet read shape: {df_fp.shape}")

os.unlink(parquet_path)

# =============================================================================
# HDF5 Files
# =============================================================================
print("\n" + "=" * 60)
print("HDF5 File Handling")
print("=" * 60)

import h5py
import tables

# Create HDF5 with h5py
with tempfile.NamedTemporaryFile(suffix=".h5", delete=False) as f:
    hdf5_path = f.name

# h5py - low level API
with h5py.File(hdf5_path, "w") as f:
    # Create datasets
    f.create_dataset("data", data=np.random.randn(1000, 100))
    f.create_dataset("labels", data=np.random.randint(0, 10, 1000))

    # Create groups
    grp = f.create_group("metadata")
    grp.attrs["version"] = "1.0"
    grp.attrs["description"] = "Sample HDF5 file"

print("Written HDF5 file with h5py")

# Read with h5py
with h5py.File(hdf5_path, "r") as f:
    print(f"\nHDF5 structure:")
    print(f"  Keys: {list(f.keys())}")
    print(f"  Data shape: {f['data'].shape}")
    print(f"  Labels shape: {f['labels'].shape}")
    print(f"  Metadata version: {f['metadata'].attrs['version']}")

os.unlink(hdf5_path)

# Pandas HDFStore
with tempfile.NamedTemporaryFile(suffix=".h5", delete=False) as f:
    hdf_pandas_path = f.name

# Write DataFrame to HDF5
df_large.to_hdf(hdf_pandas_path, key="data", mode="w", complevel=5)
print(f"\nWritten DataFrame to HDF5 (Pandas): {os.path.getsize(hdf_pandas_path) / 1024:.2f} KB")

# Read back
df_hdf = pd.read_hdf(hdf_pandas_path, key="data")
print(f"Read back shape: {df_hdf.shape}")

os.unlink(hdf_pandas_path)

# =============================================================================
# HTTP Clients
# =============================================================================
print("\n" + "=" * 60)
print("HTTP Client Libraries")
print("=" * 60)

import requests
import httpx

# Note: These examples show API usage, actual requests would need network access

print("requests library:")
print("  GET: requests.get('https://api.example.com/data')")
print("  POST: requests.post(url, json={'key': 'value'})")
print("  Headers: requests.get(url, headers={'Authorization': 'Bearer token'})")

print("\nhttpx library (async capable):")
print("  Sync: httpx.get('https://api.example.com/data')")
print("  Async: async with httpx.AsyncClient() as client:")
print("           response = await client.get(url)")

# Example of building a request (without executing)
session = requests.Session()
session.headers.update({"User-Agent": "DataScience-Notebook/1.0"})
print(f"\nSession headers: {dict(session.headers)}")

# =============================================================================
# CSV and Other Pandas I/O
# =============================================================================
print("\n" + "=" * 60)
print("Pandas I/O Capabilities")
print("=" * 60)

# Create sample data
df_sample = pd.DataFrame(
    {
        "date": pd.date_range("2024-01-01", periods=100),
        "value": np.random.randn(100),
        "category": np.random.choice(["X", "Y", "Z"], 100),
    }
)

with tempfile.TemporaryDirectory() as tmpdir:
    # CSV
    csv_path = os.path.join(tmpdir, "data.csv")
    df_sample.to_csv(csv_path, index=False)
    print(f"CSV file size: {os.path.getsize(csv_path)} bytes")

    # JSON (Pandas)
    json_path = os.path.join(tmpdir, "data.json")
    df_sample.to_json(json_path, orient="records", date_format="iso")
    print(f"JSON file size: {os.path.getsize(json_path)} bytes")

    # Feather (fast binary format)
    feather_path = os.path.join(tmpdir, "data.feather")
    df_sample.to_feather(feather_path)
    print(f"Feather file size: {os.path.getsize(feather_path)} bytes")

    # Read comparison
    df_csv = pd.read_csv(csv_path, parse_dates=["date"])
    df_json = pd.read_json(json_path)
    df_feather = pd.read_feather(feather_path)

    print(f"\nAll formats read successfully with shapes: {df_csv.shape}")

print("\n" + "=" * 60)
print("Data I/O examples complete!")
print("=" * 60)
